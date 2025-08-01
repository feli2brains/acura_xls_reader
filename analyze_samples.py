#!/usr/bin/env python3
"""
Script to analyze the Excel sample files and understand their structure.
This will help create specific plugins for each format.
"""

import os
import sys
from pathlib import Path

# Add the project root to the path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

try:
    import pandas as pd
    import openpyxl
    from openpyxl import load_workbook
except ImportError as e:
    print(f"Error importing required modules: {e}")
    print("Please install required dependencies: pip install pandas openpyxl")
    sys.exit(1)


def analyze_excel_file(file_path: str):
    """Analyze an Excel file and print its structure."""
    print(f"\n{'='*60}")
    print(f"ANALIZANDO: {file_path}")
    print(f"{'='*60}")
    
    try:
        # Load workbook
        wb = load_workbook(file_path, read_only=True)
        
        print(f"📊 Hojas encontradas: {wb.sheetnames}")
        print(f"📄 Total de hojas: {len(wb.sheetnames)}")
        
        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            print(f"\n📋 HOJA: '{sheet_name}'")
            print("-" * 40)
            
            sheet = wb[sheet_name]
            
            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            print(f"   Dimensiones: {max_row} filas x {max_col} columnas")
            
            # Sample first few rows
            print("   📝 Primeras 5 filas (primeras 10 columnas):")
            for row_idx in range(1, min(6, max_row + 1)):
                row_data = []
                for col_idx in range(1, min(11, max_col + 1)):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        row_data.append(str(cell_value)[:20])  # Truncate long values
                    else:
                        row_data.append("")
                
                print(f"      Fila {row_idx}: {row_data}")
            
            # Look for headers/keywords
            print("   🔍 Buscando palabras clave en las primeras 10 filas:")
            keywords_found = set()
            for row_idx in range(1, min(11, max_row + 1)):
                for col_idx in range(1, max_col + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value and isinstance(cell_value, str):
                        cell_lower = cell_value.lower()
                        # Look for common keywords
                        if any(keyword in cell_lower for keyword in [
                            'código', 'codigo', 'descripción', 'descripcion', 'cantidad', 
                            'precio', 'total', 'fecha', 'cliente', 'proveedor', 'parte',
                            'part number', 'item', 'material', 'unidad', 'costo'
                        ]):
                            keywords_found.add(cell_value)
            
            if keywords_found:
                print(f"      Palabras clave encontradas: {list(keywords_found)}")
            else:
                print("      No se encontraron palabras clave comunes")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error analizando {file_path}: {e}")


def main():
    """Main function to analyze all sample files."""
    samples_dir = Path("samples")
    
    if not samples_dir.exists():
        print("❌ Directorio 'samples' no encontrado")
        return
    
    excel_files = list(samples_dir.glob("*.xlsx"))
    
    if not excel_files:
        print("❌ No se encontraron archivos Excel en el directorio 'samples'")
        return
    
    print(f"🔍 Encontrados {len(excel_files)} archivos Excel para analizar:")
    for file in excel_files:
        print(f"   - {file.name}")
    
    # Analyze each file
    for file_path in excel_files:
        analyze_excel_file(str(file_path))
    
    print(f"\n{'='*60}")
    print("✅ ANÁLISIS COMPLETADO")
    print(f"{'='*60}")


if __name__ == "__main__":
    main() 