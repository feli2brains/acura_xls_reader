"""Excel utilities for the XLSX Reader Microkernel."""

import os
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import logging

from ..core.exceptions import FileNotFoundError, InvalidFileFormatError, DataProcessingError

logger = logging.getLogger(__name__)


class ExcelUtils:
    """Utility class for Excel file operations."""
    
    @staticmethod
    def validate_excel_file(file_path: str) -> bool:
        """Validate if a file is a valid Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            True if file is valid, False otherwise
        """
        try:
            if not os.path.exists(file_path):
                return False
            
            # Check file extension
            if not file_path.lower().endswith(('.xlsx', '.xls')):
                return False
            
            # Try to load the workbook
            workbook = load_workbook(file_path, read_only=True)
            workbook.close()
            return True
            
        except Exception as e:
            logger.debug(f"Excel file validation failed for {file_path}: {e}")
            return False
    
    @staticmethod
    def get_sheet_names(file_path: str) -> List[str]:
        """Get list of sheet names from Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            List of sheet names
            
        Raises:
            FileNotFoundError: If file doesn't exist
            InvalidFileFormatError: If file is not a valid Excel file
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(file_path)
            
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            
            return sheet_names
            
        except Exception as e:
            if isinstance(e, (FileNotFoundError, InvalidFileFormatError)):
                raise
            raise InvalidFileFormatError(file_path, "xlsx")
    
    @staticmethod
    def get_sheet_info(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """Get information about a sheet in an Excel file.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet (if None, gets info for all sheets)
            
        Returns:
            Dictionary with sheet information
        """
        try:
            workbook = load_workbook(file_path, read_only=True)
            
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    workbook.close()
                    raise ValueError(f"Sheet '{sheet_name}' not found")
                
                sheet = workbook[sheet_name]
                info = {
                    "name": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "dimensions": f"{sheet.max_row}x{sheet.max_column}"
                }
            else:
                info = {}
                for sheet in workbook.worksheets:
                    info[sheet.title] = {
                        "max_row": sheet.max_row,
                        "max_column": sheet.max_column,
                        "dimensions": f"{sheet.max_row}x{sheet.max_column}"
                    }
            
            workbook.close()
            return info
            
        except Exception as e:
            logger.error(f"Failed to get sheet info: {e}")
            raise DataProcessingError(f"Failed to get sheet info: {e}")
    
    @staticmethod
    def find_header_row(
        file_path: str, 
        sheet_name: str, 
        header_keywords: List[str],
        max_rows_to_check: int = 20
    ) -> Optional[int]:
        """Find the header row in an Excel sheet.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet
            header_keywords: List of keywords to look for in header
            max_rows_to_check: Maximum number of rows to check
            
        Returns:
            Row number of header (0-indexed) or None if not found
        """
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=max_rows_to_check)
            
            for row_idx, row in df.iterrows():
                row_values = [str(val).lower() for val in row.values if pd.notna(val)]
                
                # Check if any header keywords are in this row
                for keyword in header_keywords:
                    if any(keyword.lower() in val for val in row_values):
                        return row_idx
            
            return None
            
        except Exception as e:
            logger.error(f"Failed to find header row: {e}")
            return None
    
    @staticmethod
    def read_excel_with_header(
        file_path: str,
        sheet_name: str,
        header_row: int = 0,
        skip_rows: Optional[List[int]] = None
    ) -> pd.DataFrame:
        """Read Excel file with specific header row.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet
            header_row: Row number to use as header (0-indexed)
            skip_rows: Additional rows to skip
            
        Returns:
            DataFrame with the specified header
        """
        try:
            skip_rows_list = [header_row] if skip_rows is None else [header_row] + skip_rows
            
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=header_row,
                skiprows=skip_rows_list
            )
            
            # Clean column names
            df.columns = df.columns.str.strip()
            
            return df
            
        except Exception as e:
            logger.error(f"Failed to read Excel with header: {e}")
            raise DataProcessingError(f"Failed to read Excel with header: {e}")
    
    @staticmethod
    def extract_data_range(
        file_path: str,
        sheet_name: str,
        start_row: int,
        end_row: Optional[int] = None,
        start_col: int = 0,
        end_col: Optional[int] = None
    ) -> pd.DataFrame:
        """Extract a specific range of data from Excel.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet
            start_row: Starting row (0-indexed)
            end_row: Ending row (0-indexed, inclusive)
            start_col: Starting column (0-indexed)
            end_col: Ending column (0-indexed, inclusive)
            
        Returns:
            DataFrame with the extracted data
        """
        try:
            # Read the entire sheet first
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            # Extract the specified range
            if end_row is None:
                end_row = len(df)
            if end_col is None:
                end_col = len(df.columns)
            
            extracted_df = df.iloc[start_row:end_row+1, start_col:end_col+1]
            
            return extracted_df
            
        except Exception as e:
            logger.error(f"Failed to extract data range: {e}")
            raise DataProcessingError(f"Failed to extract data range: {e}")
    
    @staticmethod
    def clean_dataframe(df: pd.DataFrame, remove_empty_rows: bool = True, remove_empty_cols: bool = True) -> pd.DataFrame:
        """Clean a DataFrame by removing empty rows/columns and handling missing values.
        
        Args:
            df: DataFrame to clean
            remove_empty_rows: Whether to remove completely empty rows
            remove_empty_cols: Whether to remove completely empty columns
            
        Returns:
            Cleaned DataFrame
        """
        try:
            cleaned_df = df.copy()
            
            # Remove completely empty rows
            if remove_empty_rows:
                cleaned_df = cleaned_df.dropna(how='all')
            
            # Remove completely empty columns
            if remove_empty_cols:
                cleaned_df = cleaned_df.dropna(axis=1, how='all')
            
            # Reset index
            cleaned_df = cleaned_df.reset_index(drop=True)
            
            return cleaned_df
            
        except Exception as e:
            logger.error(f"Failed to clean DataFrame: {e}")
            raise DataProcessingError(f"Failed to clean DataFrame: {e}")
    
    @staticmethod
    def detect_data_types(df: pd.DataFrame) -> Dict[str, str]:
        """Detect data types for each column in a DataFrame.
        
        Args:
            df: DataFrame to analyze
            
        Returns:
            Dictionary mapping column names to detected data types
        """
        try:
            data_types = {}
            
            for column in df.columns:
                # Get non-null values
                non_null_values = df[column].dropna()
                
                if len(non_null_values) == 0:
                    data_types[column] = "empty"
                    continue
                
                # Try to detect type
                sample_values = non_null_values.head(100)
                
                # Check if numeric
                try:
                    pd.to_numeric(sample_values)
                    data_types[column] = "numeric"
                except (ValueError, TypeError):
                    # Check if date
                    try:
                        pd.to_datetime(sample_values, errors='raise')
                        data_types[column] = "datetime"
                    except (ValueError, TypeError):
                        # Check if boolean
                        if sample_values.dtype == bool or all(val in [True, False, 0, 1] for val in sample_values):
                            data_types[column] = "boolean"
                        else:
                            data_types[column] = "string"
            
            return data_types
            
        except Exception as e:
            logger.error(f"Failed to detect data types: {e}")
            return {}
    
    @staticmethod
    def validate_dataframe_schema(df: pd.DataFrame, expected_schema: Dict[str, Any]) -> List[str]:
        """Validate DataFrame against an expected schema.
        
        Args:
            df: DataFrame to validate
            expected_schema: Expected schema dictionary
            
        Returns:
            List of validation error messages (empty if valid)
        """
        errors = []
        
        try:
            # Check required columns
            required_columns = expected_schema.get("required_columns", [])
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                errors.append(f"Missing required columns: {missing_columns}")
            
            # Check data types
            expected_types = expected_schema.get("column_types", {})
            detected_types = ExcelUtils.detect_data_types(df)
            
            for column, expected_type in expected_types.items():
                if column in df.columns:
                    detected_type = detected_types.get(column, "unknown")
                    if detected_type != expected_type:
                        errors.append(
                            f"Column '{column}' has type '{detected_type}' but expected '{expected_type}'"
                        )
            
            # Check value constraints
            constraints = expected_schema.get("constraints", {})
            for column, constraint in constraints.items():
                if column in df.columns:
                    if "min" in constraint:
                        min_val = constraint["min"]
                        if df[column].min() < min_val:
                            errors.append(f"Column '{column}' has values below minimum {min_val}")
                    
                    if "max" in constraint:
                        max_val = constraint["max"]
                        if df[column].max() > max_val:
                            errors.append(f"Column '{column}' has values above maximum {max_val}")
                    
                    if "unique" in constraint and constraint["unique"]:
                        if df[column].duplicated().any():
                            errors.append(f"Column '{column}' has duplicate values but should be unique")
            
            return errors
            
        except Exception as e:
            errors.append(f"Schema validation failed: {e}")
            return errors 